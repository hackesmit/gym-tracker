"""Min-Max Program (.xlsx) layout: parser, format detection, sheet resolution,
and the /import-program upload path.

The synthetic workbook mirrors the real "2726Min-Max_Program_5x.xlsx" layout
cell-for-cell for the parts the parser depends on: blank column A, a title
banner, "Block 1" / "Intro Week" label rows, a "Week N" header row followed by
two sub-header rows (Set 1/Set 2, Load/Reps), session labels only on the first
exercise row of a session, "Rest Day" rows, per-set load/reps tracking cells
that carry the owner's personal log, RIR per set instead of RPE, and rep /
warm-up ranges that Excel mangled into datetimes.
"""

from __future__ import annotations

import datetime
import io
import os

import openpyxl
import pytest

from app.models import Program, ProgramExercise
from app.parser import (
    _rir_to_rpe,
    candidate_sheet_names,
    detect_sheet_format,
    parse_program,
    parse_workbook,
    resolve_sheet,
    resolve_sheet_name,
)

HEADER = [
    None, "Week {n}", "Exercise", "Last-Set Intensity Technique", "Warm-up Sets",
    "Working Sets", "Rep Range", "Tracking Load and Reps", None, None, None,
    "Failure?", None, "Rest", "Substitution Option 1", "Substitution Option 2", "Notes",
]
SUB1 = [None] * 7 + ["Set 1", None, "Set 2", None]
SUB2 = [None] * 7 + ["Load", "Reps", "Load", "Reps", "RIR (Set 1)", "RIR (Set 2)"]

D = datetime.datetime


def _ex(session, name, tech, warm, sets, reps, load1, reps1, load2, reps2, rir1, rir2, rest, s1, s2, notes):
    return [None, session, name, tech, warm, sets, reps, load1, reps1, load2, reps2, rir1, rir2, rest, s1, s2, notes]


def _week_rows(n: int, tech: str = "N/A", rir=(1, 0)):
    return [
        [h.format(n=n) if isinstance(h, str) else h for h in HEADER],
        SUB1,
        SUB2,
        _ex("Upper 1", "Barbell Incline Press", "N/A", D(2025, 2, 4), 2, D(2025, 6, 8), 185, 8, None, 6, 2, 1, "3-5 min", "Smith Machine Incline Press", "DB Incline Press", "Pause at the bottom."),
        _ex(None, "Pec Deck", tech, D(2025, 1, 2), 2, D(2025, 6, 8), 175, 8, 175, 7, rir[0], rir[1], "1-2 min", "DB Flye", "Cable Flye", "Squeeze."),
        _ex(None, "Dead Hang (optional)", "N/A", "0-1", 2, "N/A", None, None, None, None, 0, 0, "1-2 min", "N/A", "N/A", "Add a few seconds each week."),
        _ex("Lower 1", "Lying Leg Curl", "N/A", D(2025, 1, 2), 2, D(2025, 8, 10), 140, 10, 140, 9, 1, 0, "1-2 min", "Seated Leg Curl", "Nordic Ham Curl", "Big stretch."),
        _ex(None, "Machine Hip Abduction", "N/A", "0-1", 1, D(2025, 6, 8), "max", 11, None, None, 0, "N/A", "1-2 min", "Cable Hip Abduction", "Standing Plate Abduction", None),
        [None, "Rest Day"],
        _ex("Upper 2", "Close-Grip Lat Pulldown", "N/A", D(2025, 2, 3), 2, D(2025, 8, 10), 200, 11, 200, 8, 2, 1, "2-3 min", "Close-Grip Pull-Up", "1-Arm Cable Pulldown", "Lean back 15 degrees."),
        _ex("Lower 2", "Barbell RDL", "N/A", D(2025, 2, 3), 2, D(2025, 6, 8), 225, 8, 225, 8, 3, 2, "2-3 min", "DB RDL", "Seated Cable Deadlift", "Hips back."),
        _ex("Arms/Delts", "Bayesian Cable Curl", "N/A", "0-1", 2, D(2025, 6, 8), None, None, None, None, 1, 0, "1-2 min", "Incline DB Curl", "Standing DB Curl", "Control the negative."),
        [None, "Rest Day"],
        [],
    ]


def build_minmax_workbook(sheet_title: str = "5x Per Week", weeks: int = 2) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    rows = [
        [None, "The Min-Max Program"],
        [None] + [None] * 15 + ["Copyright 2025 by Jeff Nippard. All rights reserved. "],
        [None, "IMPORTANT PROGRAM NOTES (PLEASE READ BEFORE STARTING!)\n\nSome bullet points here."],
        [],
        [None] * 6 + ["WARM-UP PROTOCOL"],
        [None] * 6 + ["Start each workout with about 5 minutes of light cardio."],
        [],
        [None, "Block 1"],
        [None, "Intro Week"],
    ]
    for w in range(1, weeks + 1):
        # Week 2+ uses an intensity technique on Pec Deck and RIR 0/0
        rows += _week_rows(w) if w == 1 else _week_rows(w, tech="Myo-reps", rir=(0, 0))
        if w == 1:
            rows += [[None, "Block 2"], [None, "Deload Week"]] if weeks > 2 else []
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture()
def minmax_path(tmp_path):
    p = tmp_path / "minmax.xlsx"
    p.write_bytes(build_minmax_workbook())
    return p


def test_detect_format_minmax(minmax_path):
    wb = openpyxl.load_workbook(minmax_path, data_only=True)
    assert detect_sheet_format(wb["5x Per Week"]) == "minmax"


def test_resolve_sheet_name_variants():
    assert resolve_sheet_name(["2x Week", "3x Week", "4x Week", "5x Week"], 4) == "4x Week"
    assert resolve_sheet_name(["5x Per Week"], 5) == "5x Per Week"
    assert resolve_sheet_name(["Only Sheet"], 3) == "Only Sheet"
    assert resolve_sheet_name(["3x Per Week", "5x Per Week"], 4) is None
    # "15x" must not match a 5x request
    assert resolve_sheet_name(["15x Per Week", "Notes"], 5) is None
    # known names outrank fuzzy matches regardless of workbook order
    assert candidate_sheet_names(["5x Overview", "5x Per Week", "5x Week"], 5) == [
        "5x Week", "5x Per Week", "5x Overview",
    ]


def test_resolve_sheet_skips_frequency_named_sheet_without_program(tmp_path):
    """An "Overview" tab that happens to be named 5x must not shadow the real
    program sheet that comes after it in the workbook."""
    wb = openpyxl.Workbook()
    ov = wb.active
    ov.title = "5x Overview"
    ov.append(["Read the PDF first"])
    real = wb.create_sheet("5x Per Week")
    real.append([None, "The Min-Max Program"])
    for r in _week_rows(1):
        real.append(r)
    assert resolve_sheet(wb, 5) == "5x Per Week"
    p = tmp_path / "ov.xlsx"
    wb.save(p)
    parsed = parse_workbook(p, 5)
    assert parsed["sheet_name"] == "5x Per Week" and len(parsed["exercises"]) == 8


def test_detect_format_requires_minmax_header_conjunction():
    """An Essentials header with an incidental Min-Max word stays Essentials."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["WEEK 1", "EXERCISE", "WARM-UP SETS", "WORKING SETS", "REPS", "LOAD", "RPE", "REST", "SUB 1", "SUB 2", "NOTES (RIR guidance / Failure?)"])
    assert detect_sheet_format(ws) == "essentials"
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.append([None, "Week 1", "Exercise", "Warm-up Sets", "Working Sets", "Rep Range", "Failure?", "Rest"])
    assert detect_sheet_format(ws2) == "minmax"
    wb3 = openpyxl.Workbook()
    ws3 = wb3.active
    ws3.append(["nothing here"])
    assert detect_sheet_format(ws3) is None


def test_header_row_beyond_sixty_rows_is_found(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "5x Per Week"
    ws.append([None, "The Min-Max Program"])
    for _ in range(80):
        ws.append([None, "cover material"])
    for r in _week_rows(1):
        ws.append(r)
    p = tmp_path / "long.xlsx"
    wb.save(p)
    parsed = parse_workbook(p, 5)
    assert parsed["format"] == "minmax" and len(parsed["exercises"]) == 8


def test_rir_to_rpe_decimals_ranges_and_annotations():
    assert _rir_to_rpe(["2", "1"]) == "8-9"
    assert _rir_to_rpe(["1.5", "N/A"]) == "8.5"
    assert _rir_to_rpe(["1-2", None]) == "8-9"
    assert _rir_to_rpe(["2 RIR", "0"]) == "8-10"
    assert _rir_to_rpe(["N/A", "", None]) == ""
    assert _rir_to_rpe([0]) == "10"


def test_week_with_extra_set_columns_remaps_headers(tmp_path):
    """Week 2 adds a third tracking set (Load/Reps) plus a third RIR column, so
    Failure? / Rest / substitutions / notes shift right by three. Rows must be
    read with week 2's mapping, and no tracking value may leak into a field."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "5x Per Week"
    ws.append([None, "The Min-Max Program"])
    for r in _week_rows(1):
        ws.append(r)
    hdr2 = [
        None, "Week 2", "Exercise", "Last-Set Intensity Technique", "Warm-up Sets", "Working Sets",
        "Rep Range", "Tracking Load and Reps", None, None, None, None, None,
        "Failure?", None, None, "Rest", "Substitution Option 1", "Substitution Option 2", "Notes",
    ]
    ws.append(hdr2)
    ws.append([None] * 7 + ["Set 1", None, "Set 2", None, "Set 3", None])
    ws.append([None] * 7 + ["Load", "Reps", "Load", "Reps", "Load", "Reps", "RIR (Set 1)", "RIR (Set 2)", "RIR (Set 3)"])
    ws.append([None, "Upper 1", "Pec Deck", "Myo-reps", "0-1", 3, D(2025, 8, 10), 175, 9, 175, 8, 175, 7, 1, 1, 0, "1-2 min", "DB Flye", "Cable Flye", "Week two notes."])
    p = tmp_path / "shift.xlsx"
    wb.save(p)
    ex = parse_workbook(p, 5)["exercises"]
    w2 = [e for e in ex if e["week"] == 2]
    assert len(w2) == 1
    e = w2[0]
    assert e["working_sets"] == 3
    assert e["prescribed_reps"] == "8-10"
    assert e["prescribed_rpe"] == "9-10"
    assert e["rest_period"] == "1-2 min"
    assert e["substitution_1"] == "DB Flye" and e["substitution_2"] == "Cable Flye"
    assert e["notes"] == "Last set: Myo-reps. Week two notes."
    for v in e.values():
        assert v not in (175, 9, 8, 7)
    # week 1 rows still parse with their own mapping
    assert [x for x in ex if x["week"] == 1][0]["rest_period"] == "3-5 min"


def test_parse_minmax_structure(minmax_path):
    parsed = parse_workbook(minmax_path, 5)
    assert parsed["format"] == "minmax"
    assert parsed["sheet_name"] == "5x Per Week"
    assert parsed["title"] == "The Min-Max Program"
    assert parsed["detected_frequency"] == 5
    ex = parsed["exercises"]
    assert len(ex) == 2 * 8

    weeks = sorted({e["week"] for e in ex})
    assert weeks == [1, 2]
    w1 = [e for e in ex if e["week"] == 1]
    sessions = []
    for e in w1:
        if e["session_name"] not in sessions:
            sessions.append(e["session_name"])
    assert sessions == ["UPPER 1", "LOWER 1", "UPPER 2", "LOWER 2", "ARMS/DELTS"]
    assert [e["session_order_in_week"] for e in w1] == [1, 1, 1, 2, 2, 3, 4, 5]
    assert [e["exercise_order"] for e in w1 if e["session_name"] == "UPPER 1"] == [1, 2, 3]

    # (week, session, order) is unique: satisfies uq_program_exercise
    keys = [(e["week"], e["session_name"], e["exercise_order"]) for e in ex]
    assert len(keys) == len(set(keys))


def test_parse_minmax_fields(minmax_path):
    ex = parse_workbook(minmax_path, 5)["exercises"]
    by = {(e["week"], e["exercise_name_canonical"]): e for e in ex}

    incline = by[(1, "BARBELL INCLINE PRESS")]
    assert incline["exercise_name_raw"] == "Barbell Incline Press"
    assert incline["warm_up_sets"] == "2-4"          # datetime(2025,2,4) un-mangled
    assert incline["working_sets"] == 2
    assert incline["prescribed_reps"] == "6-8"       # datetime(2025,6,8) un-mangled
    assert incline["prescribed_rpe"] == "8-9"        # RIR 2 / 1
    assert incline["rest_period"] == "3-5 min"
    assert incline["substitution_1"] == "Smith Machine Incline Press"
    assert incline["substitution_2"] == "DB Incline Press"
    assert incline["notes"] == "Pause at the bottom."

    rdl = by[(1, "BARBELL RDL")]
    assert rdl["prescribed_rpe"] == "7-8"            # RIR 3 / 2
    assert rdl["prescribed_reps"] == "6-8"

    abd = by[(1, "MACHINE HIP ABDUCTION")]
    assert abd["working_sets"] == 1
    assert abd["prescribed_rpe"] == "10"             # RIR 0 / N/A
    assert abd["notes"] is None

    hang = by[(1, "DEAD HANG (OPTIONAL)")]
    assert hang["prescribed_reps"] == ""             # N/A rep range
    assert hang["substitution_1"] is None and hang["substitution_2"] is None
    assert hang["prescribed_rpe"] == "10"

    pec2 = by[(2, "PEC DECK")]
    assert pec2["prescribed_rpe"] == "10"
    assert pec2["notes"] == "Last set: Myo-reps. Squeeze."
    pec1 = by[(1, "PEC DECK")]
    assert pec1["notes"] == "Squeeze."               # N/A technique is not folded in


def test_parse_minmax_never_exports_personal_loads(minmax_path):
    ex = parse_workbook(minmax_path, 5)["exercises"]
    for e in ex:
        assert not {"load", "load_kg", "reps_completed"} & set(e)
        # the sheet's tracked loads (185, 175, 140, 225, "max") must not leak into any field
        for v in e.values():
            assert v not in (185, 175, 140, 225, 200) and v != "max"


def test_session_label_on_its_own_row_starts_a_session(tmp_path):
    """Variant layout: session name on a row of its own, exercises beneath."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "3x Per Week"
    ws.append([None, "The Min-Max Program"])
    ws.append([h.format(n=1) if isinstance(h, str) else h for h in HEADER])
    ws.append(SUB1)
    ws.append(SUB2)
    ws.append([None, "Full Body 1"])
    ws.append(_ex(None, "Pec Deck", "N/A", "0-1", 2, D(2025, 6, 8), None, None, None, None, 1, 0, "1-2 min", "N/A", "N/A", None))
    ws.append(_ex(None, "Leg Extension", "N/A", "0-1", 2, D(2025, 6, 8), None, None, None, None, 1, 0, "1-2 min", "N/A", "N/A", None))
    ws.append([None, "Rest Day"])
    ws.append([None, "Full Body 2"])
    ws.append(_ex(None, "Leg Press", "N/A", "0-1", 2, D(2025, 6, 8), None, None, None, None, 1, 0, "1-2 min", "N/A", "N/A", None))
    p = tmp_path / "v.xlsx"
    wb.save(p)
    ex = parse_workbook(p, 3)["exercises"]
    assert [(e["session_name"], e["session_order_in_week"], e["exercise_order"]) for e in ex] == [
        ("FULL BODY 1", 1, 1), ("FULL BODY 1", 1, 2), ("FULL BODY 2", 2, 1),
    ]


def test_parse_program_legacy_entry_point_autodetects(minmax_path):
    ex = parse_program(minmax_path, "5x Per Week")
    assert len(ex) == 16 and ex[0]["exercise_name_canonical"] == "BARBELL INCLINE PRESS"


def test_import_program_endpoint_accepts_minmax(client, db):
    data = build_minmax_workbook()
    r = client.post(
        "/api/import-program",
        files={"file": ("2726Min-Max_Program_5x.xlsx", data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"frequency": "5"},
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["format"] == "minmax"
    assert body["frequency"] == 5
    assert body["program_name"] == "The Min-Max Program"
    assert body["total_exercises"] == 16
    assert body["sessions_per_week"]["1"] and len(body["sessions_per_week"]["1"]) == 5

    prog = db.query(Program).filter(Program.id == body["program_id"]).one()
    assert prog.status == "active"
    assert prog.frequency == 5
    assert prog.total_weeks == 12
    assert prog.source_file == "2726Min-Max_Program_5x.xlsx"
    n = db.query(ProgramExercise).filter(ProgramExercise.program_id == prog.id).count()
    assert n == 16


def test_import_program_single_sheet_wrong_frequency_uses_detected(client, db):
    """A single-sheet workbook is used regardless of the picked frequency and
    the stored frequency reflects the sessions actually found."""
    data = build_minmax_workbook(sheet_title="Program")
    r = client.post(
        "/api/import-program",
        files={"file": ("mm.xlsx", data, "application/octet-stream")},
        data={"frequency": "3", "program_name": "My Min-Max"},
    )
    assert r.status_code == 200, r.text
    assert r.json()["frequency"] == 5
    assert r.json()["program_name"] == "My Min-Max"


def test_import_program_unsupported_session_count_is_422(client):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Program"
    ws.append([None, "Solo"])
    ws.append([h.format(n=1) if isinstance(h, str) else h for h in HEADER])
    ws.append(_ex("Only Day", "Pec Deck", "N/A", "0-1", 2, D(2025, 6, 8), None, None, None, None, 1, 0, "1-2 min", "N/A", "N/A", None))
    buf = io.BytesIO()
    wb.save(buf)
    r = client.post(
        "/api/import-program",
        files={"file": ("solo.xlsx", buf.getvalue(), "application/octet-stream")},
        data={"frequency": "2"},
    )
    assert r.status_code == 422
    assert "1 training sessions per week" in r.json()["detail"]


def test_import_program_no_matching_sheet_is_422(client):
    wb = openpyxl.Workbook()
    wb.active.title = "3x Per Week"
    wb.create_sheet("4x Per Week")
    buf = io.BytesIO()
    wb.save(buf)
    r = client.post(
        "/api/import-program",
        files={"file": ("mm.xlsx", buf.getvalue(), "application/octet-stream")},
        data={"frequency": "5"},
    )
    assert r.status_code == 422
    assert "No sheet for 5x/week" in r.json()["detail"]


REAL_FILE = os.environ.get("MINMAX_XLSX")


@pytest.mark.skipif(not REAL_FILE or not os.path.exists(REAL_FILE or ""), reason="MINMAX_XLSX not set")
def test_real_minmax_file_parses_full_program():
    parsed = parse_workbook(REAL_FILE, 5)
    ex = parsed["exercises"]
    assert parsed["format"] == "minmax"
    assert parsed["detected_frequency"] == 5
    assert len(ex) == 420
    assert sorted({e["week"] for e in ex}) == list(range(1, 13))
    for w in range(1, 13):
        assert len({e["session_name"] for e in ex if e["week"] == w}) == 5
    assert all(e["working_sets"] in (1, 2) for e in ex)
